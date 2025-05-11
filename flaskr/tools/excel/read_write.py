import os
import pandas as pd
from openpyxl import load_workbook

# Paths
current_path = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(current_path, "filled_report.xlsx")
output_path = os.path.join(current_path, "filled_report_with_revenue.xlsx")

# 1) Read Excel, starting from row where table begins (assume starts at row 5, column C)
# Skip the header rows manually, and read from correct columns
df = pd.read_excel(file_path, usecols="A:C", skiprows=4, nrows=5, header=None) # nrows to read
df.columns = ['name', 'qty', 'price']

# 2) Calculate revenue = 10% of price
df['revenue'] = df['price'] * 0.10

# 3) Load workbook and write back using openpyxl (preserve structure, start at column "D")
wb = load_workbook(file_path)
ws = wb.active

# Write headers
ws["D3"] = "Revenue(10%)"

# Write revenue values
for i, value in enumerate(df['revenue'], start=5):  # Start from row 5
    ws[f"D{i}"] = value

# 4) Save to a new file
wb.save(output_path)
print("✓ Saved → filled_report_with_revenue.xlsx")
